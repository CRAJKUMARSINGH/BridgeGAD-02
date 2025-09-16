import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

def create_bridge_sample_excel():
    """Create comprehensive Excel sample input file for Bridge applications"""
    
    # Bridge parameters from LISP and Python analysis
    bridge_params = {
        'Parameter': [
            'LBRIDGE', 'NSPAN', 'TOPRL', 'SOFL', 'LEFT', 'ABTLEN', 'ALFL', 'ARFL',
            'SCALE1', 'SCALE2', 'CCBR', 'SKEW', 'KERBW', 'KERBD', 'SLBTHC', 
            'SLBTHE', 'SLBTHT', 'PIERTW', 'BATTR', 'PIERST', 'SPAN1', 'SPAN2',
            'FUTD', 'FUTW', 'FUTL', 'DWTH', 'ALCW', 'ALCD', 'ALFB', 'ALFBL',
            'ALTB', 'ALTBL', 'ALFO', 'ALFD', 'ALBB', 'ALBBL', 'CAPT', 'CAPB',
            'CAPW', 'PIERN', 'RTL', 'DATUM'
        ],
        'Value': [
            30000, 1, 110000, 108000, 0, 10000, 105000, 105000,
            100, 1, 50, 0, 1000, 500, 300,
            250, 150, 1500, 12, 10000, 15000, 15000,
            1000, 3000, 5000, 500, 2000, 500, 6000, 101500,
            12000, 100500, 1000, 1000, 6000, 101500, 103000, 102500,
            2000, 1, 105000, 100000
        ],
        'Unit': [
            'mm', 'nos', 'mm', 'mm', 'mm', 'mm', 'mm', 'mm',
            '-', '-', 'mm', 'degrees', 'mm', 'mm', 'mm',
            'mm', 'mm', 'mm', 'ratio', 'mm', 'mm', 'mm',
            'mm', 'mm', 'mm', 'mm', 'mm', 'mm', 'mm', 'mm',
            'mm', 'mm', 'mm', 'mm', 'mm', 'mm', 'mm', 'mm',
            'mm', 'nos', 'mm', 'mm'
        ],
        'Description': [
            'Bridge Length', 'Number of Spans', 'Top Road Level', 'Soffit Level', 
            'Left Chainage', 'Abutment Length', 'Left Abutment Floor Level', 'Right Abutment Floor Level',
            'Plan/Elevation Scale', 'Section Scale', 'Carriage Width Between Rails', 'Skew Angle',
            'Kerb Width', 'Kerb Depth', 'Slab Thickness at Center', 'Slab Thickness at Edge',
            'Slab Thickness at Toe', 'Pier Top Width', 'Batter Ratio', 'Pier Stem Height',
            'Span 1 Length', 'Span 2 Length', 'Foundation Depth', 'Foundation Width',
            'Foundation Length', 'Dirt Wall Thickness', 'Approach Left Course Width',
            'Approach Left Course Depth', 'Approach Left Foundation Breadth', 'Approach Left Foundation Bottom Level',
            'Approach Left Top Breadth', 'Approach Left Top Bottom Level', 'Approach Left Foundation Offset',
            'Approach Left Foundation Depth', 'Approach Left Back Breadth', 'Approach Left Back Bottom Level',
            'Cap Top Level', 'Cap Bottom Level', 'Cap Width', 'Number of Piers', 'Return Level', 'Datum Level'
        ]
    }
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Bridge Parameters"
    
    # Get data from dictionary
    params = bridge_params['Parameter']
    values = bridge_params['Value']
    units = bridge_params['Unit']
    descriptions = bridge_params['Description']
    
    # Headers
    headers = ['Parameter', 'Value', 'Unit', 'Description']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row, (param, value, unit, desc) in enumerate(zip(params, values, units, descriptions), 2):
        ws.cell(row=row, column=1, value=param).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=unit)
        ws.cell(row=row, column=4, value=desc)
    
    # Auto-fit columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

def create_multi_span_sample():
    """Create multi-span bridge sample"""
    bridge_params = {
        'Parameter': [
            'LBRIDGE', 'NSPAN', 'TOPRL', 'SOFL', 'LEFT', 'SPAN1', 'SPAN2', 'SPAN3',
            'PIERN', 'CCBR', 'SKEW', 'KERBW', 'KERBD', 'SCALE1', 'SCALE2'
        ],
        'Value': [
            60000, 3, 110000, 108000, 0, 20000, 20000, 20000,
            2, 50, 15, 1000, 500, 100, 1
        ],
        'Description': [
            'Total Bridge Length', 'Number of Spans', 'Top Road Level', 'Soffit Level', 
            'Left Chainage', 'Span 1 Length', 'Span 2 Length', 'Span 3 Length',
            'Number of Piers', 'Carriage Width', 'Skew Angle', 'Kerb Width', 'Kerb Depth',
            'Plan Scale', 'Section Scale'
        ]
    }
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Multi-Span Bridge"
    
    # Headers
    headers = ['Parameter', 'Value', 'Description']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
    
    # Data
    for row, (param, value, desc) in enumerate(zip(bridge_params['Parameter'], bridge_params['Value'], bridge_params['Description']), 2):
        ws.cell(row=row, column=1, value=param).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=desc)
    
    return wb

def create_skew_bridge_sample():
    """Create skew bridge sample"""
    bridge_params = {
        'Parameter': ['LBRIDGE', 'NSPAN', 'TOPRL', 'SOFL', 'SKEW', 'CCBR', 'SPAN1'],
        'Value': [25000, 1, 110000, 108000, 30, 50, 25000],
        'Description': ['Bridge Length', 'Number of Spans', 'Top Level', 'Soffit Level', 
                       'Skew Angle (30 degrees)', 'Carriage Width', 'Span Length']
    }
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Skew Bridge"
    
    # Headers with different color
    headers = ['Parameter', 'Value', 'Description']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")
    
    # Data
    for row, (param, value, desc) in enumerate(zip(bridge_params['Parameter'], bridge_params['Value'], bridge_params['Description']), 2):
        ws.cell(row=row, column=1, value=param).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=desc)
    
    return wb

if __name__ == "__main__":
    # Create output directory
    output_dir = "C:/Users/Rajkumar/BridgeGAD-02/SAMPLE_INPUT_FILES"
    os.makedirs(output_dir, exist_ok=True)
    
    # Create sample files
    print("Creating Bridge Sample Excel Files...")
    
    # Standard bridge
    wb1 = create_bridge_sample_excel()
    wb1.save(f"{output_dir}/bridge_standard_input.xlsx")
    print("✓ Standard bridge sample created")
    
    # Multi-span bridge
    wb2 = create_multi_span_sample()
    wb2.save(f"{output_dir}/bridge_multispan_input.xlsx")
    print("✓ Multi-span bridge sample created")
    
    # Skew bridge
    wb3 = create_skew_bridge_sample()
    wb3.save(f"{output_dir}/bridge_skew_input.xlsx")
    print("✓ Skew bridge sample created")
    
    print(f"\\nAll sample files created in: {output_dir}")
    print("Files created:")
    print("- bridge_standard_input.xlsx")
    print("- bridge_multispan_input.xlsx") 
    print("- bridge_skew_input.xlsx")